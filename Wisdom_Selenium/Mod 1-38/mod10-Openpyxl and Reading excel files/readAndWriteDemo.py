

import openpyxl

path = "C:\\Users\\Desktop\\test.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active

print(sheet.cell(7, 3).value)
print("Total no. of columns : " + str(sheet.max_column))
print("Total no. of rows : " + str(sheet.max_row))

w = sheet.cell(4, 2)
w.value = "Simran"
wb.save(path)
print("data updated")

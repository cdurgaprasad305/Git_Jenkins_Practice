import requests
from selenium import webdriver
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import time

serv_obj=Service("C:\Drivers\chromedriver_win32\chromedriver.exe")
driver=webdriver.Chrome(service=serv_obj)

driver.get("https://demo.nopcommerce.com/")
driver.maximize_window()

# click on link
driver.find_element(By.LINK_TEXT,"Digital downloads").click()
driver.find_element(By.PARTIAL_LINK_TEXT,"Digital").click()

#Find number of links in a page
links=driver.find_elements(By.TAG_NAME,'a')
link=driver.find_elements(By.XPATH,'//a')
print("total number of links:",len(links))

# print all the link names
for link in links:
    print(link.text)

# --------------------------------------------------------

serv_obj=Service("C:\Drivers\chromedriver_win32\chromedriver.exe")
driver=webdriver.Chrome(service=serv_obj)

driver.get("https://www.opencart.com/index.php?route=account/register")
driver.maximize_window()

# drpcountry_ele=driver.find_element(By.XPATH,"//select[@id='input-country']")
drop_country=Select(driver.find_element(By.XPATH,"//select[@id='input-country']"))

#select option from the dropdown
drop_country.select_by_visible_text("India")
drop_country.select_by_value("10")  #Argentina
drop_country.select_by_index(13)  # index

# capture all the options and print them
all_options=drop_country.options
print("total number of options:",len(all_options))

for opt in all_options:
     print(opt.text)

# select option from dropdown without using built-in method
for opt in all_options:
     if opt.text=="India":
         opt.click()
         break

allOptions=driver.find_elements(By.XPATH,'//*[@id="input-country"]/option')
print(len(allOptions))

# ---------------**** Checkbox ****-------------------------------------------


serv_obj=Service("C:\Drivers\chromedriver_win32\chromedriver.exe")
driver=webdriver.Chrome(service=serv_obj)

driver.get("https://itera-qa.azurewebsites.net/home/automation")
driver.maximize_window()

# 1) select specific checkbox
driver.find_element(By.XPATH,"//input[@id='monday']").click()

# 2) select all the checkboxes
checkboxes=driver.find_elements(By.XPATH,"//input[@type='checkbox' and contains(@id,'day')]")
print(len(checkboxes)) #7

#Appraoch1
for i in range(len(checkboxes)):
     checkboxes[i].click()

#Appraoch2
for checkbox in checkboxes:
    checkbox.click()


# 3) select multiple checkboxes by choice
for checkbox in checkboxes:
     weekname=checkbox.get_attribute('id')
     if weekname=='monday' or weekname=='sunday':
         checkbox.click()

# 4 ) select last 2 checkboxes
# range(5,7) -->6,7
# totalnumberofelements-2=starting index
for i in range(len(checkboxes)-2,len(checkboxes)):
    checkboxes[i].click()

# 5) select first 2 checkboxes
for i in range(len(checkboxes)):
    if i<2:
        checkboxes[i].click()

# 6) clearing all the checkboxes
for checkbox in checkboxes:
    if checkbox.is_selected():
        checkbox.click()
driver.quit()
# --------------------***** Links **** -----------------------------------------
# we need to install requests package through File-->Settings-->ProjectIntrepreter-->+--> requests


serv_obj=Service("C:\Drivers\chromedriver_win32\chromedriver.exe")
driver=webdriver.Chrome(service=serv_obj)
# Total number of broken links
driver.get("http://www.deadlinkcity.com/")
driver.maximize_window()

allLinks=driver.find_elements(By.TAG_NAME,'a')
count=0

for link in allLinks:
    url=link.get_attribute('href')
    try:
        res=requests.head(url)
    except:
        None

    if res.status_code>=400:
        print(url,"  is broken link")
        count+=1
    else:
        print(url,"   is valid link")

print("Total number of broken links:",count)

# -------------------**** Alert ***** ---------------------------------------------
# Alert Windows

driver.get("https://the-internet.herokuapp.com/javascript_alerts")
driver.maximize_window()

#opens alert window
driver.find_element(By.XPATH,"//button[normalize-space()='Click for JS Prompt']").click()
time.sleep(5)

alert_window=driver.switch_to.alert

print(alert_window.text)
alert_window.send_keys("welcome")


#alert_window.accept() #close alert window by using OK button
alert_window.dismiss() #close alert window by using Cancel button)

driver.get("https://mypage.rediff.com/login/dologin")
driver.maximize_window()

driver.find_element(By.XPATH,"//input[@value='Login']").click() #lOGIN BUTTON
time.sleep(5)
driver.switch_to.alert.accept()

driver.close()


#driver.get("http://the-internet.herokuapp.com/basic_auth")
driver.get("http://admin:admin@the-internet.herokuapp.com/basic_auth")

driver.maximize_window()

# -----------------**** Browser Windows Handle **** -----------------------------------
# Handle Browser Windows

driver.get("https://opensource-demo.orangehrmlive.com/")
driver.maximize_window()

windowid=driver.current_window_handle
print(windowid) #CDwindow-24CFB4FFB7CB60C7309293A9217B9F2A
                #CDwindow-E7283DA0F55EF107A63C1CD349434080

driver.find_element(By.LINK_TEXT,"OrangeHRM, Inc").click()
windowsIDs=driver.window_handles

#Approach1
parentwindowid=windowsIDs[0] #CDwindow-A82AD2C060A4BBCAFAD3DF3C6172CCBE
childwindowid=windowsIDs[1] #CDwindow-F63E2A3D2324F7132FFD60C9A8E16A95
print(parentwindowid,childwindowid)

driver.switch_to.window(childwindowid)
print("title of the child window",driver.title)

driver.switch_to.window(parentwindowid)
print("title of the parent window",driver.title)

#Approach2

for winid in windowsIDs:
    driver.switch_to.window(winid)
    print(driver.title)

time.sleep(3)

# Close specific browser windows   1 2 3
for winid in windowsIDs:
    driver.switch_to.window(winid)
    if driver.title=="OrangeHRM HR Software | Free HR Software | HRMS | HRIS" or driver.title=='XYZ':
        driver.close()

# Multiple Windows

driver.get("https://testautomationpractice.blogspot.com/")

driver.find_element(By.XPATH,"//input[@id='Wikipedia1_wikipedia-search-input']").send_keys("testing")
driver.find_element(By.XPATH,"//input[@type='submit']").click()

searchlinks=driver.find_elements(By.XPATH,"//div[@id='Wikipedia1_wikipedia-search-results']//div/a")
print("Number of links:",len(searchlinks))

print("printing and clicking on links...............")
for link in searchlinks:
    print(link.text)
    link.click()

#After opening all the pages, capture windowid's
windowIds=driver.window_handles

print("Switching to each browser window and getting the titles........")
for windowid in windowIds:
    driver.switch_to.window(windowid)
    print(driver.title)

driver.quit()


# -----------------------**** Frames **** ------------------------------------------

# Frames

driver.get("https://www.selenium.dev/selenium/docs/api/java/index.html?overview-summary.html")
driver.maximize_window()

driver.switch_to.frame("packageListFrame")
driver.find_element(By.LINK_TEXT,"org.openqa.selenium").click()
driver.switch_to.default_content()  # go back to main page

driver.switch_to.frame("packageFrame")
driver.find_element(By.LINK_TEXT,"WebDriver").click()
driver.switch_to.default_content()  # go back to main page

driver.switch_to.frame("classFrame")
driver.find_element(By.XPATH,"/html/body/header/nav/div[1]/div[1]/ul/li[8]").click()

# Inner Frames
driver.find_element(By.XPATH,"//a[normalize-space()='Iframe with in an Iframe']").click()

outerframe=driver.find_element(By.XPATH,"//iframe[@src='MultipleFrames.html']")
driver.switch_to.frame(outerframe)

innerframe=driver.find_element(By.XPATH,"/html/body/section/div/div/iframe")
driver.switch_to.frame(innerframe)

driver.find_element(By.XPATH,"//input[@type='text']").send_keys("welcome")

driver.switch_to.parent_frame()  # directly switch to parent frame(outerframe)

# --------------------------*****  Date Time ***** ---------------------------
# Date Time
import time

d1='02/11/2021'   # DD/MM/YYYY
d2='05/12/2021'  #DD/MM/YYYY

dep_date = time.strptime(d1, "%d/%m/%Y")
return_date = time.strptime(d2, "%d/%m/%Y")

print(return_date>dep_date)  # returns true/false

# true means return date is greater than departure date
# false means return date is Not greater than departure date

# ---------------------------**** ActionChains *****-------------------------------
# Double click

driver.get("https://www.w3schools.com/tags/tryit.asp?filename=tryhtml5_ev_ondblclick3")
driver.maximize_window()

driver.switch_to.frame("iframeResult") # switch to frame

field1=driver.find_element(By.XPATH,"//input[@id='field1']")
field1.clear()
field1.send_keys("welcome")

button=driver.find_element(By.XPATH,"//button[normalize-space()='Copy Text']")

act=ActionChains(driver)
act.double_click(button).perform() # double click action

# --------------------------------------------------------------------------------
# Drag and Drop
driver.get("http://www.dhtmlgoodies.com/scripts/drag-drop-custom/demo-drag-drop-3.html")
driver.maximize_window()

act=ActionChains(driver)

rome_ele=driver.find_element(By.ID,"box6")
italy_ele=driver.find_element(By.ID,"box106")
act.drag_and_drop(rome_ele,italy_ele).perform()  # drag and drop opration

wsh_ele=driver.find_element(By.ID,"box3")
italy_ele=driver.find_element(By.ID,"box103")
act.drag_and_drop(wsh_ele,italy_ele).perform()  # drag and drop opration


# --------------------------------------------------------------------------------
# Mouse Hover *

driver.get("https://opensource-demo.orangehrmlive.com")
driver.maximize_window()

#Login
driver.find_element(By.ID,"txtUsername").send_keys("Admin")
driver.find_element(By.ID,"txtPassword").send_keys("admin123")
driver.find_element(By.ID,"btnLogin").click()
time.sleep(3)

#Admin-->user management-->users
admin=driver.find_element(By.XPATH,"//*[@id='menu_admin_viewAdminModule']/b")
usermgmt=driver.find_element(By.XPATH,"//*[@id='menu_admin_UserManagement']")
users=driver.find_element(By.XPATH,"//*[@id='menu_admin_viewSystemUsers']")

act=ActionChains(driver)
act.move_to_element(admin).move_to_element(usermgmt).move_to_element(users).click().perform()

# --------------------------------------------------------------------------------

# Right Click
driver.get("http://swisnl.github.io/jQuery-contextMenu/demo.html")
driver.maximize_window()

button=driver.find_element(By.XPATH,"//span[@class='context-menu-one btn btn-neutral']")

act=ActionChains(driver)
act.context_click(button).perform()  # right click

time.sleep(5)
driver.find_element(By.XPATH,"//span[normalize-space()='Copy']").click() # copy
time.sleep(3)
driver.switch_to.alert.accept()


#Keyboard Actions

driver.get("https://text-compare.com/")
driver.maximize_window()

input1=driver.find_element(By.XPATH,"//textarea[@id='inputText1']")
input2=driver.find_element(By.XPATH,"//textarea[@id='inputText2']")

input1.send_keys("welcome to selenium")

act=ActionChains(driver)

#input1 ---> Ctrl+A  Select the text
act.key_down(Keys.CONTROL)
act.send_keys("a")
act.key_up(Keys.CONTROL)
act.perform()
act.key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()

#input1  --->Ctrl+C  Copy text
act.key_down(Keys.CONTROL)
act.send_keys("c")
act.key_up(Keys.CONTROL)
act.perform()
act.key_down(Keys.CONTROL).send_keys("c").key_up(Keys.CONTROL).perform()

# Press Tab key to navigate to input2(second)
act.send_keys(Keys.TAB)
act.perform()
act.send_keys(Keys.TAB).perform()

#input2 --->Ctrl+V Past the text
act.key_down(Keys.CONTROL)
act.send_keys("v")
act.key_up(Keys.CONTROL)
act.perform()
act.key_down(Keys.CONTROL).send_keys("v").key_up(Keys.CONTROL).perform()


# -------------------------- **** Scrolling ***** -------------------------------------
# Scrolling

driver.get("https://www.countries-ofthe-world.com/flags-of-the-world.html")
driver.maximize_window()

# 1. Scroll down page by pixel
driver.execute_script("window.scrollBy(0,3000)","")
value=driver.execute_script("return window.pageYOffset;")
print("Number of pixels moved:",value) #3000

# 2. Scroll down page till the element is visible
flag=driver.find_element(By.XPATH,"//img[@alt='Flag of India']")
driver.execute_script("arguments[0].scrollIntoView();",flag)
value=driver.execute_script("return window.pageYOffset;")
print("Number of pixels moved:",value) #5038.3330078125

# 3.scroll down page till end
driver.execute_script("window.scrollBy(0,document.body.scrollHeight)")
value=driver.execute_script("return window.pageYOffset;")
print("Number of pixels moved:",value) #5990.8330078125

time.sleep(5)
# Scroll up to starting position
driver.execute_script("window.scrollBy(0,-document.body.scrollHeight)")
value=driver.execute_script("return window.pageYOffset;")
print("Number of pixels moved:",value) #0

# -------------------------**** Slider ****** ---------------------------------
# Slider
driver.get("https://www.jqueryscript.net/demo/Price-Range-Slider-jQuery-UI/")
driver.maximize_window()

min_slider=driver.find_element(By.XPATH,"//body//div//span[1]")
max_slider=driver.find_element(By.XPATH,"//body//div//span[2]")

print("Location of sliders Before moving........")
print(min_slider.location) #    {'x': 59, 'y': 252}
print(max_slider.location) #    {'x': 639, 'y': 252}


act=ActionChains(driver)

act.drag_and_drop_by_offset(min_slider,100,0).perform()
act.drag_and_drop_by_offset(max_slider,-39,0).perform()

print("Location of sliders After moving........")
print(min_slider.location) # {'x': 158, 'y': 252}
print(max_slider.location) # {'x': 598, 'y': 252}


# ---------------------- **** Upload & Download File **** -------------------------------------
# Upload File
driver.get("https://www.monsterindia.com/")
driver.maximize_window()

driver.find_element(By.XPATH,"//span[@class='uprcse semi-bold']").click()
driver.find_element(By.XPATH,"//*[@id='file-upload']").send_keys("C:\SeleniumPractice\sample.pdf")

# --------------------------------------------------------------------------------
# File Download
import os
location=os.getcwd()

def chrome_setup():
    from selenium.webdriver.chrome.service import Service
    serv_obj = Service("C:\Drivers\chromedriver_win32\chromedriver.exe")
    #download files in desired location
    preferences={"download.default_directory":location}
    ops=webdriver.ChromeOptions()
    ops.add_experimental_option("prefs",preferences)
    driver = webdriver.Chrome(service=serv_obj,options=ops)
    return driver


#Automation code
driver=chrome_setup()

driver.get("https://file-examples.com/index.php/sample-documents-download/sample-doc-download/")
driver.maximize_window()
driver.find_element(By.XPATH,"//tbody/tr[1]/td[5]/a[1]").click()

# --------------------- **** Tabs & Windows  ******  ------------------------------------
# Tabs & Windows

driver.get("https://demo.nopcommerce.com/")
driver.maximize_window()
regilink=Keys.CONTROL+Keys.RETURN
driver.find_element(By.LINK_TEXT,"Register").send_keys(regilink)


#New Tab - Selenium4 : Opens a new tab and switches to new tab
driver.get("https://www.opencart.com/")
driver.switch_to.new_window('tab')
driver.get("https://www.orangehrm.com/")

#New Window - Selenium4 :  Opens a new browser window and switches to new window
driver.get("https://www.opencart.com/")
driver.switch_to.new_window('window')
driver.get("https://www.orangehrm.com/")



# ------------------**** Headless testing  ******  -------------------------------
# Headless testing

def headless_chrome():
    from selenium.webdriver.chrome.service import Service
    serv_obj = Service("C:\Drivers\chromedriver_win32\chromedriver.exe")
    ops=webdriver.ChromeOptions()
    ops.headless=True
    driver=webdriver.Chrome(service=serv_obj,options=ops)
    return driver

driver=headless_chrome()
#driver=headless_edge()
# driver=headless_firefox()


driver.get("https://demo.nopcommerce.com/")
print(driver.title)
print(driver.current_url)
driver.close()

# --------------- **** Capture Screenshot ***** --------------------------------
# Capture Screenshot

driver.get("https://demo.nopcommerce.com/")
driver.maximize_window()

driver.save_screenshot("C:\\Users\\admin\\PycharmProjects\\PythonSelenium\\day23\\homepage.png")
driver.save_screenshot(os.getcwd()+"\\homepage.png")
driver.get_screenshot_as_file(os.getcwd()+"\\homepage.png")

driver.get_screenshot_as_png()  #driver.get_screenshot_as_base64() #saves in binary format

driver.quit()


# -------------------*****   DB Operations  *****   ------------------------------------
# DB Operations

import mysql.connector

try:
    con=mysql.connector.connect(host="localhost",port=3306,user="root",passwd="root",database="mydb")
    curs=con.cursor()  # create curosor
    curs.execute("select * from orders")  #execute query through cursor

    for row in curs:
        print(row[0],row[1],row[2])

    con.close()  # close connection
except:
    print("Connection unsuccessful...")

print("Finished.....")

# --------------------------------------------------------------------------------


insert_query="insert into student values(104,'Kim')"
update_query="update student set sname='Mary' where sid=104"
delete_query="delete from student where sid=104"

import mysql.connector

try:
    con=mysql.connector.connect(host="localhost",port=3307,user="root",passwd="root",database="mydb")
    curs=con.cursor()  # create cursor
    curs.execute(delete_query)  #execute query through cursor
    con.commit() # commit transaction
    con.close()
except:
    print("Connection unsuccessful...")

print("Finished.....")


# --------------------------------------------------------------------------------


driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html")
driver.maximize_window()

try:
    con = mysql.connector.connect(host="localhost", port=3306, user="root", passwd="root", database="mydb")
    curs = con.cursor()  # create cursor
    curs.execute("select * from caldata")  # execute query through cursor

    for row in curs:
        # reading data from excel
        pric =row[0]
        rateofinterest =row[1]
        per1 =row[2]
        per2 =row[3]
        fre =row[4]
        exp_mvalue =row[5]
        # passing data to the application
        driver.find_element(By.XPATH, "//input[@id='principal']").send_keys(pric)
        driver.find_element(By.XPATH, "//input[@id='interest']").send_keys(rateofinterest)
        driver.find_element(By.XPATH, "//input[@id='tenure']").send_keys(per1)
        perioddrp = Select(driver.find_element(By.XPATH, "//select[@id='tenurePeriod']"))
        perioddrp.select_by_visible_text(per2)
        frequencydrp = Select(driver.find_element(By.XPATH, "//select[@id='frequency']"))
        frequencydrp.select_by_visible_text(fre)
        driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[1]/img").click()  # calculate button

        act_mvalue = driver.find_element(By.XPATH, "//span[@id='resp_matval']/strong").text

        # Validation
        if float(exp_mvalue) == float(act_mvalue):
            print("test passed")
        else:
            print("test failed")
        driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[2]/img").click()
        time.sleep(2)
    con.close()
except:
    print("Connection unsuccessful....")

driver.close()


# -----------------**** Excel Operations  ****** -------------------------------
# Excel Operations
import openpyxl

#File-->Workbook--->Sheets-->Rows-->Cells

file="C:\SeleniumPractice\data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row # count number of rows in a excel sheet 6
cols=sheet.max_column # count number of columns in a excel sheet  4

#Reading all the rows & columns from excel sheet
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='       ')
    print()


# --------------------------------------------------------------------------------

file="C:\\SeleniumPractice\\testdata.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook.active   #  (or) sheet=workbook["Data"]     --  get active sheet from excel
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="welcome"
workbook.save(file)

#multiple data
file="C:\\SeleniumPractice\\testdata1.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook.active   #  (or) sheet=workbook["Data"]     --  get active sheet from excel

sheet.cell(1,1).value=123
sheet.cell(1,2).value="smith"
sheet.cell(1,3).value="engineer"

sheet.cell(2,1).value=567
sheet.cell(2,2).value="john"
sheet.cell(2,3).value="manager"

sheet.cell(3,1).value=567
sheet.cell(3,2).value="david"
sheet.cell(3,3).value="developer"

workbook.save(file)   # save the file after entering the data

# --------------------------------------------------------------------------------

driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html")
driver.maximize_window()

file="C:\SeleniumPractice\caldata.xlsx"
rows=XLUtils.getRowCount(file,"Sheet1")

for r in range(2,rows+1):
    #reading data from excel
    pric=XLUtils.readData(file,"Sheet1",r,1)
    rateofinterest=XLUtils.readData(file,"Sheet1",r,2)
    per1 = XLUtils.readData(file, "Sheet1", r, 3)
    per2 = XLUtils.readData(file, "Sheet1", r, 4)
    fre = XLUtils.readData(file, "Sheet1", r, 5)
    exp_mvalue = XLUtils.readData(file, "Sheet1", r, 6)

    #passing data to the application
    driver.find_element(By.XPATH,"//input[@id='principal']").send_keys(pric)
    driver.find_element(By.XPATH,"//input[@id='interest']").send_keys(rateofinterest)
    driver.find_element(By.XPATH,"//input[@id='tenure']").send_keys(per1)
    perioddrp=Select(driver.find_element(By.XPATH,"//select[@id='tenurePeriod']"))
    perioddrp.select_by_visible_text(per2)
    frequencydrp=Select(driver.find_element(By.XPATH,"//select[@id='frequency']"))
    frequencydrp.select_by_visible_text(fre)
    driver.find_element(By.XPATH,"//*[@id='fdMatVal']/div[2]/a[1]/img").click()  # calculate button

    act_mvalue=driver.find_element(By.XPATH,"//span[@id='resp_matval']/strong").text

    #Validation
    if float(exp_mvalue)==float(act_mvalue):
        print("test passed")
        XLUtils.writeData(file,"Sheet1",r,8,"Passed")
        XLUtils.fillGreenColor(file,"Sheet1",r,8)
    else:
        print("test failed")
        XLUtils.writeData(file,"Sheet1",r,8,"Failed")
        XLUtils.fillRedColor(file,"Sheet1",r,8)
    driver.find_element(By.XPATH,"//*[@id='fdMatVal']/div[2]/a[2]/img").click()
    time.sleep(2)

driver.close()

# --------------------------------------------------------------------------------

import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)

def fillGreenColor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                       end_color='60b212',
                       fill_type='solid')
    sheet.cell(rownum,columnno).fill=greenFill
    workbook.save(file)


def fillRedColor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                       end_color='ff0000',
                       fill_type='solid')
    sheet.cell(rownum,columnno).fill=redFill
    workbook.save(file)
# --------------------------------------------------------------------------------

# Mouse Drag and Drop

driver.get("https://testautomationpractice.blogspot.com/")
driver.maximize_window()

item1 = driver.find_element(By.XPATH,"//ul[@id='gallery']/li[1]")
item2 = driver.find_element(By.XPATH,"//ul[@id='gallery']/li[2]")

trash = driver.find_element(By.XPATH,"//div[@id='trash']")

act=ActionChains(driver)

act.drag_and_drop(item1, trash).perform()
act.drag_and_drop(item2, trash).perform()

# --------------------------------------------------------------------------------

driver.get("https://www.cit.com/cit-bank/resources/calculators/certificate-of-deposit-calculator/")
driver.maximize_window()

inideposit = driver.find_element(By.XPATH,"//input[@id='mat-input-0']")
length = driver.find_element(By.XPATH,"//input[@id='mat-input-1']")
apr = driver.find_element(By.XPATH,"//input[@id='mat-input-2']")
calbutton = driver.find_element(By.XPATH,"//button[@id='CIT-chart-submit']")

path = "C:\\SeleniumPractice\\caldata2.xlsx"
rows = XLUtils.getRowCount(path, "Sheet1")
print("row count is : " , rows)

for r in range(2,rows+1):
    inidepo = XLUtils.readData(path, "Sheet1", r, 1)
    interestrate = XLUtils.readData(path, "Sheet1", r, 2)
    monthlength = XLUtils.readData(path, "Sheet1", r, 3)
    compoundingmonths = XLUtils.readData(path, "Sheet1", r, 4)
    exptotal = XLUtils.readData(path, "Sheet1", r, 5)

    inideposit.clear()
    length.clear()
    apr.clear()
    time.sleep(3)

    inideposit.send_keys(inidepo)
    length.send_keys(monthlength)
    apr.send_keys(interestrate)

    #Dropdown (Boostrap) - Not having Select Tag
    compoundrp = driver.find_element(By.XPATH,"//mat-select[@id='mat-select-0']")
    compoundrp.click()
    options = driver.find_elements(By.XPATH,"//div[@id='mat-select-0-panel']//mat-option")

    for option in options:
        if(option.text==compoundingmonths):
            option.click()

    calbutton.click()
    acttotal = driver.find_element(By.XPATH,"//span[@id='displayTotalValue']").text

    print("exp total is from excel: ", exptotal)
    print("act total is from app: ", acttotal)

    if(exptotal==acttotal):
        XLUtils.writeData(path, "Sheet1" ,r, 7 ,"Passed")
        XLUtils.fillGreenColor(path, "Sheet1" ,r, 7)
    else:
        XLUtils.writeData(path, "Sheet1" ,r, 7 ,"Failed")
        XLUtils.fillRedColor(path, "Sheet1" ,r, 7)

print("calculation has been completed")
driver.close()

# --------------------------------------------------------------------------------
#Test Case
#------------------
# 1) Open Web Browser(Chrome/firefox/Edge).
# 2) Open URL  https://opensource-demo.orangehrmlive.com/
# 3) Enter username  (Admin).
# 4) Enter password  (admin123).
# 5) Click on Login.
# 6) Capture title of the home page.(Actual title)
# 7) Verify title of the page: OrangeHRM    (Expected)
# 8) close browser

from selenium import webdriver

# Selenium 3
driver=webdriver.Chrome(executable_path="C:\Drivers\chromedriver_win32\chromedriver.exe")
    ## driver=webdriver.Chrome()
driver.get("https://opensource-demo.orangehrmlive.com/")
driver.find_element_by_name("txtUsername").send_keys("Admin")
driver.find_element_by_id("txtPassword").send_keys("admin123")
driver.find_element_by_id("btnLogin").click()
act_title=driver.title
exp_title="OrangeHRM"
if act_title==exp_title:
    print("Login Test Passed")
else:
    print("Login Test Failed")
driver.close()

# Selenium 4
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
serv_obj=Service("C:\Drivers\chromedriver_win32\chromedriver.exe")
driver=webdriver.Chrome(service=serv_obj)
##driver=webdriver.Chrome()
driver.get("https://opensource-demo.orangehrmlive.com/")
driver.find_element(By.NAME,"txtUsername").send_keys("Admin")
driver.find_element(By.ID,"txtPassword").send_keys("admin123")
driver.find_element(By.ID,"btnLogin").click()


act_title=driver.title
exp_title="OrangeHRM"
if act_title==exp_title:
    print("Login Test Passed")
else:
    print("Login Test Failed")
driver.close()

# ------------------------------  LOCATORS  ------------------------------------------


driver.get("https://demo.nopcommerce.com/")
driver.maximize_window()  #maximize the browser window

# id & name locators
driver.find_element(By.ID,"small-searchterms").send_keys("Lenovo Thinkpad X1 Carbon Laptop")
driver.find_element(By.NAME,"q").send_keys("Lenovo Thinkpad X1 Carbon Laptop")


# LINK_TEXT & PARTIAL_LINK_TEXT
driver.find_element(By.LINK_TEXT,"Register").click()
driver.find_element(By.PARTIAL_LINK_TEXT,"Reg").click()


sliders=driver.find_elements(By.CLASS_NAME,'homeslider-container')
print(len(sliders)) #5 total number of sliders on home page

links=driver.find_elements(By.TAG_NAME,'a')
print(len(links)) #  149 total number of links on home page

# ------------------------  CSS_SELECTOR
#tag & id
driver.find_element(By.CSS_SELECTOR,"input#email").send_keys("abc")
driver.find_element(By.CSS_SELECTOR,"#email").send_keys("abc")

# tag and class
driver.find_element(By.CSS_SELECTOR,"input.inputtext").send_keys("abc@gmail.com")
driver.find_element(By.CSS_SELECTOR,".inputtext").send_keys("abc@gmail.com")

# tag & attribute
driver.find_element(By.CSS_SELECTOR,"input[data-testid=royal_email]").send_keys("abc@gmail.com")
driver.find_element(By.CSS_SELECTOR,"[data-testid=royal_email]").send_keys("abc@gmail.com")


# tag , class & attribute
driver.find_element(By.CSS_SELECTOR,"input.inputtext[data-testid=royal_pass]").send_keys("xyz")

# ------------------------  XPATH
#Absulte xpath
driver.find_element(By.XPATH,"/html/body/div/div[1]/header/div[3]/div/div/div[2]/form/input[4]").send_keys("T-shirts")
driver.find_element(By.XPATH,"/html/body/div/div[1]/header/div[3]/div/div/div[2]/form/button").click()

#Relative xpath
driver.find_element(By.XPATH,"//*[@id='search_query_top']").send_keys("T-shirts")
driver.find_element(By.XPATH,"//button[@name='submit_search']").click()

#or  and
driver.find_element(By.XPATH,"//input[@id='search_query_top' or @name='search_query' ]").send_keys("T-shirts")
driver.find_element(By.XPATH,"//button[@name='submit_search' and @class='btn btn-default button-search']").click()

#contains()   & start-with()
driver.find_element(By.XPATH,"//input[contains(@id,'search')]").send_keys("T-shirts")
driver.find_element(By.XPATH,"//button[starts-with(@name,'submit_')]").click()

#text()
driver.find_element(By.XPATH,"//a[text()='Women']").click()

#self

text_msg=driver.find_element(By.XPATH,"//a[contains(text(),'India Tourism De')]/self::a").text
print(text_msg) #India Tourism De

#parent
text_msg=driver.find_element(By.XPATH,"//a[contains(text(),'India Tourism De')]/parent::td").text
print(text_msg) #India Tourism De

#child
childs=driver.find_elements(By.XPATH,"//a[contains(text(),'India Tourism De')]/ancestor::tr/child::td")
print(len(childs))  #5

#Ancestor
text_msg=driver.find_element(By.XPATH,"//a[contains(text(),'India Tourism De')]/ancestor::tr").text
print(text_msg) #India Tourism De A 358.35 375.30 + 4.73

#Decendant
descendants=driver.find_elements(By.XPATH,"//a[contains(text(),'India Tourism De')]/ancestor::tr/descendant::*")
print("Number of descendant nodes:",len(descendants)) #7

#Following
followings=driver.find_elements(By.XPATH,"//a[contains(text(),'India Tourism De')]/ancestor::tr/following::*")
print("Number of descendant nodes:",len(followings)) #719

#Folowing-sibling
followingsiblings=driver.find_elements(By.XPATH,"//a[contains(text(),'India Tourism De')]/ancestor::tr/following-sibling::*")
print("Number of descendant nodes:",len(followingsiblings)) #72

#preceding
precedings=driver.find_elements(By.XPATH,"//a[contains(text(),'India Tourism De')]/ancestor::tr/preceding::*")
print(len(precedings)) #251

#preceding-sibling
precedingsiblings=driver.find_elements(By.XPATH,"//a[contains(text(),'India Tourism De')]/ancestor::tr/preceding-sibling::tr")
print(len(precedingsiblings)) #11

# --------------------------------------------------------------------------------

#total rows n a table
rows=len(driver.find_elements(By.XPATH,"//table[@id='resultTable']//tbody/tr"))
print("total number of rows in a table:",rows)

count=0
for r in range(1,rows+1):
    status=driver.find_element(By.XPATH,"//table[@id='resultTable']/tbody/tr["+str(r)+"]/td[5]").text
    if status=="Enabled":
        count=count+1

print("Total Number of users:",rows)
print("Number of enabled users:",count)
print("Number of disabled users:",(rows-count))

# ---------------- ****    Date Picker   **** -----------------------------------------
# Date Picker

year="2022"
month="June"
date="10"

driver.find_element(By.XPATH,"//*[@id='datepicker']").click() # opens datepicker

while True:
    mon=driver.find_element(By.XPATH,"//span[@class='ui-datepicker-month']").text
    yr=driver.find_element(By.XPATH,"//span[@class='ui-datepicker-year']").text

    if mon==month and yr==year:
        break;
    else:
        driver.find_element(By.XPATH,"//*[@id='ui-datepicker-div']/div/a[2]/span").click()  #Next arrow
        #driver.find_element(By.XPATH,"//*[@id='ui-datepicker-div']/div/a[1]/span").click() # Previous Arrow - old date

#select date

dates=driver.find_elements(By.XPATH,"//div[@id='ui-datepicker-div']//table/tbody/tr/td/a")

for ele in dates:
    if ele.text==date:
        ele.click()
        break


# -------------------- ****  Dynamic Wait **** -------------------------------

url = "https://testautomationpractice.blogspot.com/"
driver.maximize_window()
driver.implicitly_wait(10)
driver.get(url)

# Find the element by ID
# This is a dynamic element, so we will wait for it to be displayed
# The element is a span with ID "Stats1_totalCount"

tot_count = driver.find_element(By.ID,"Stats1_totalCount")
# Dynamic wait for the element to be displayed
for i in range(1,20):
    if tot_count.is_displayed():
        break
    else:
        time.sleep(0.5)
        print("web element is not displayed yet, waiting..." ,i)

print("Total Visit Count:",tot_count.get_property("textContent"))
# print(tot_count.get_property("innerText"))
# print(tot_count.get_property("innerHTML"))

driver.quit()

# -------------------------------------------------------------------------------

